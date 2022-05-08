from io import StringIO
from openpyxl import Workbook
from xlsxq import __version__
from xlsxq import RangeShowQuery
from xlsxq import SheetListQuery
import json
import pytest
import xlsxq


def test_version():
    assert __version__ == '0.1.3'


def test_parse_arguments_exits_with_help_without_args():
    with pytest.raises(SystemExit) as pytest_wrapped_e:
        xlsxq.parse_arguments([])
    assert pytest_wrapped_e.type == SystemExit
    assert pytest_wrapped_e.value.code != 0


def test_sheet_list_query_prints_sheet_list(tmpdir):
    book_path = tmpdir.join("book.xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws.title == 'Sheet'
    ws['A1'] = 1
    wb.save(str(book_path))
    query = SheetListQuery(infile=str(book_path), output='json')
    io = StringIO()
    query.execute(file=io)
    data = json.loads(io.getvalue())
    assert data == [{'name': 'Sheet'}]


def test_range_show_query_prints_values(tmpdir):
    book_path = tmpdir.join("book.xlsx")
    wb = Workbook()
    ws = wb.create_sheet("Mysheet")
    ws['A1'] = 11
    ws['B1'] = 12
    ws['A2'] = 21
    ws['B2'] = 22
    ws['A3'] = 31
    ws['B3'] = 32
    wb.save(str(book_path))
    query = RangeShowQuery(
            infile=str(book_path), sheet='Mysheet', range_='A1:B3',
            output='tsv')
    io = StringIO()
    query.execute(file=io)
    text = io.getvalue()
    lines = text.splitlines()
    assert len(lines) == 3
    assert lines[0] == "11\t12"
    assert lines[1] == "21\t22"
    assert lines[2] == "31\t32"


def test_range_show_query_prints_values_when_range_is_cell(tmpdir):
    book_path = tmpdir.join("book.xlsx")
    wb = Workbook()
    ws = wb.create_sheet("Mysheet")
    ws['A1'] = 11
    wb.save(str(book_path))
    query = RangeShowQuery(
            infile=str(book_path), sheet='Mysheet', range_='A1', output='json')
    io = StringIO()
    io = StringIO()
    query.execute(file=io)
    data = json.loads(io.getvalue())
    assert data == [[11]]


def test_range_show_query_prints_values_when_range_is_row(tmpdir):
    book_path = tmpdir.join("book.xlsx")
    wb = Workbook()
    ws = wb.create_sheet("Mysheet")
    ws['A1'] = 11
    ws['B1'] = 12
    wb.save(str(book_path))
    query = RangeShowQuery(
            infile=str(book_path), sheet='Mysheet', range_='A1:B1',
            output='json')
    io = StringIO()
    io = StringIO()
    query.execute(file=io)
    data = json.loads(io.getvalue())
    assert data == [[11, 12]]


def test_range_show_query_prints_values_when_range_is_col(tmpdir):
    book_path = tmpdir.join("book.xlsx")
    wb = Workbook()
    ws = wb.create_sheet("Mysheet")
    ws['A1'] = 11
    ws['A2'] = 21
    wb.save(str(book_path))
    query = RangeShowQuery(
            infile=str(book_path), sheet='Mysheet', range_='A1:A2',
            output='json')
    io = StringIO()
    io = StringIO()
    query.execute(file=io)
    data = json.loads(io.getvalue())
    assert data == [[11], [21]]
